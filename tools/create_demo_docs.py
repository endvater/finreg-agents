#!/usr/bin/env python3
"""
FinRegAgents – Demo-Dokumenten-Generator

Erzeugt synthetische Prüfungsunterlagen für Musterbank AG,
die einen realistischen GwG-Prüffall abbilden – mit bewussten Lücken und
Mängeln, um alle vier Bewertungsstufen (konform / teilkonform /
nicht_konform / nicht_prüfbar) im Bericht sichtbar zu machen.

Verwendung:
    python tools/create_demo_docs.py
    python tools/create_demo_docs.py --output ./demo
"""

import argparse
import json
import logging
import random
import sys
from datetime import date, timedelta
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ------------------------------------------------------------------ #
# Abhängigkeiten prüfen
# ------------------------------------------------------------------ #

try:
    from fpdf import FPDF
except ImportError:
    sys.exit(
        "fpdf2 fehlt. Bitte installieren: pip install fpdf2\n"
        "Oder: pip install -r requirements.txt"
    )

try:
    import pandas as pd
    import openpyxl  # noqa: F401
except ImportError:
    sys.exit("pandas / openpyxl fehlt. Bitte installieren: pip install pandas openpyxl")


# ------------------------------------------------------------------ #
# Text-Sanitizer (FPDF core fonts = latin-1 only)
# ------------------------------------------------------------------ #


def _s(text: str) -> str:
    """Ersetze Nicht-Latin-1-Zeichen durch sichere ASCII-Äquivalente."""
    replacements = {
        "\u2013": "-",  # En-Dash
        "\u2014": "--",  # Em-Dash
        "\u2018": "'",  # Geschweifte Anführung links
        "\u2019": "'",  # Geschweifte Anführung rechts
        "\u201c": '"',  # Geschweifte Anführung links (doppelt)
        "\u201d": '"',  # Geschweifte Anführung rechts (doppelt)
        "\u2022": "-",  # Bullet
        "\u00b7": "-",  # Mittelpunkt
        "\u2026": "...",  # Ellipsis
    }
    for ch, repl in replacements.items():
        text = text.replace(ch, repl)
    return text.encode("latin-1", errors="replace").decode("latin-1")


# ------------------------------------------------------------------ #
# PDF-Basisklasse
# ------------------------------------------------------------------ #


class BankPDF(FPDF):
    """FPDF-Unterklasse mit Musterbank-Header/Footer."""

    BANK = "Musterbank AG"
    VERTRAULICH = "VERTRAULICH - NUR FUER INTERNE ZWECKE"

    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.cell(0, 6, self.BANK, align="L")
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 6, self.VERTRAULICH, align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(180, 180, 180)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(3)

    def footer(self):
        self.set_y(-13)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 6, f"Seite {self.page_no()} | {self.BANK}", align="C")
        self.set_text_color(0, 0, 0)


def _pdf_new(orientation="P") -> BankPDF:
    pdf = BankPDF(orientation=orientation, unit="mm", format="A4")
    pdf.set_margins(left=20, top=20, right=20)
    pdf.set_auto_page_break(auto=True, margin=18)
    return pdf


def _h1(pdf: BankPDF, text: str):
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_fill_color(230, 240, 255)
    pdf.cell(0, 10, _s(text), fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)


def _h2(pdf: BankPDF, text: str):
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(30, 70, 140)
    pdf.cell(0, 8, _s(text), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.set_draw_color(30, 70, 140)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(3)


def _body(pdf: BankPDF, text: str):
    pdf.set_font("Helvetica", size=10)
    pdf.multi_cell(0, 6, _s(text), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)


def _label_value(pdf: BankPDF, label: str, value: str):
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(65, 7, _s(label) + ":", new_x="RIGHT", new_y="TOP")
    pdf.set_font("Helvetica", size=10)
    pdf.multi_cell(0, 7, _s(value), new_x="LMARGIN", new_y="NEXT")


def _table_header(pdf: BankPDF, cols: list[tuple[str, int]]):
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(50, 90, 160)
    pdf.set_text_color(255, 255, 255)
    for label, width in cols:
        pdf.cell(width, 7, _s(label), border=1, fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)


def _table_row(pdf: BankPDF, values: list[str], widths: list[int], alt: bool = False):
    pdf.set_font("Helvetica", size=9)
    if alt:
        pdf.set_fill_color(240, 245, 255)
    else:
        pdf.set_fill_color(255, 255, 255)
    for val, width in zip(values, widths):
        pdf.cell(width, 6, _s(val), border=1, fill=True)
    pdf.ln()


# ------------------------------------------------------------------ #
# PDF-Dokumente
# ------------------------------------------------------------------ #


def create_gwb_bestellungsurkunde(out: Path):
    """
    S04-01 → KONFORM: Bestellung ordnungsgemäß, BaFin gemeldet.
    S04-02 → TEILKONFORM: Nur 1 Stellvertreter für 520 Mitarbeiter.
    S04-03 → KONFORM: Berichtspflicht geregelt.
    """
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "Bestellungsurkunde – Geldwäschebeauftragter")

    _label_value(pdf, "Dokument-Nr.", "MB-GWB-2022-001")
    _label_value(pdf, "Version", "2.1 (gültig ab 15.03.2022)")
    _label_value(pdf, "Erstellt von", "Vorstand / Compliancefunktion")
    _label_value(pdf, "Genehmigt am", "14.03.2022 – Vorstandsbeschluss")
    _label_value(pdf, "BaFin-Meldung", "16.03.2022, Aktenzeichen 2022/GWB/0471")
    pdf.ln(4)

    _h2(pdf, "1. Bestellung und Funktion")
    _body(
        pdf,
        "Die Musterbank AG bestellt hiermit Herrn Dr. Thomas Müller, geboren 12. April 1974, "
        "zum Geldwäschebeauftragten (GwB) gemäß § 7 Abs. 1 GwG. "
        "Die Bestellung gilt ab dem 15. März 2022 auf unbestimmte Zeit.\n\n"
        "Der GwB ist der Geschäftsleitung unmittelbar unterstellt und berichtet monatlich "
        "sowie anlassbezogen an den Gesamtvorstand. Er ist mit allen notwendigen Befugnissen "
        "ausgestattet, um seiner gesetzlichen Aufgabe nachzukommen (§ 7 Abs. 5 GwG), "
        "insbesondere hat er Zugang zu sämtlichen Daten, Dokumenten und Systemen.",
    )

    _h2(pdf, "2. Qualifikation")
    _body(
        pdf,
        "Dr. Müller verfügt über folgende für die Tätigkeit als GwB relevante Qualifikationen:\n"
        "– Studium der Rechtswissenschaften (Erstes und Zweites Staatsexamen, Universität Frankfurt)\n"
        "– 12 Jahre Berufserfahrung im Bankenbereich, davon 8 Jahre im Compliance-Umfeld\n"
        "– Zertifizierung CAMS (Certified Anti-Money Laundering Specialist) seit 2017, rezertifiziert 2020, 2023\n"
        "– Teilnahme an jährlichen BaFin-GwG-Schulungen (zuletzt November 2024)\n"
        "– Mitglied der ACAMS Deutschland-Sektion seit 2018",
    )

    _h2(pdf, "3. Stellvertretung")
    _body(
        pdf,
        "Als Stellvertreter wird Frau Sandra Becker, Compliance Officer, bestellt "
        "(Bestellungsdatum: 15.03.2022, BaFin-Meldung: 16.03.2022). "
        "Frau Becker ist zertifizierte Compliance-Fachkraft (ICA) und verfügt über "
        "5 Jahre Berufserfahrung in der GwG-Compliance.\n\n"
        "Hinweis: Das Team des GwB umfasst aktuell 2 Personen (GwB + 1 Stellvertreter) "
        "für die gesamte Bank mit 520 Mitarbeitern und einem Bestandskundenportfolio "
        "von ca. 38.000 Kunden. Eine Evaluierung der personellen Ressourcenausstattung "
        "ist für Q2 2025 geplant, aber noch nicht abgeschlossen.",
    )

    _h2(pdf, "4. Berichtspflichten")
    _body(
        pdf,
        "Der GwB erstattet der Geschäftsleitung gemäß § 7 Abs. 5 GwG und den internen "
        "Regelungen folgende Berichte:\n"
        "– Monatlicher AML-Statusbericht (Vorlage an Gesamtvorstand, 1. Vorstandssitzung im Folgemonat)\n"
        "– Jahresbericht AML (bis 31. Januar des Folgejahres, mit GL-Genehmigung)\n"
        "– Ad-hoc-Berichte bei wesentlichen Ereignissen (Verdachtsmeldungen, BaFin-Anfragen)\n\n"
        "Der Jahresbericht 2024 wurde am 28. Januar 2025 vom Gesamtvorstand genehmigt "
        "(Protokoll Vorstandssitzung 28.01.2025, TOP 3).",
    )

    _h2(pdf, "5. Weisungsfreiheit und Schutz")
    _body(
        pdf,
        "Der GwB ist bei der Wahrnehmung seiner Aufgaben weisungsfrei. "
        "Er kann ohne seine Zustimmung nur aus wichtigem Grund abberufen werden (§ 7 Abs. 7 GwG). "
        "Ihm steht ein direktes Vortragsrecht beim Vorstandsvorsitzenden zu.",
    )

    pdf.output(out)
    log.info("  ✓ %s", out.name)


def create_risikoanalyse(out: Path):
    """
    S01-01 → KONFORM: Risikoanalyse liegt vor.
    S01-02 → NICHT_KONFORM: Letzte vollständige Aktualisierung Feb 2022 – älter als 12 Monate.
    S01-03 → TEILKONFORM: Vertriebswege nur oberflächlich abgedeckt.
    S01-04 → TEILKONFORM: GL-Genehmigung nicht klar im Dokument nachweisbar.
    """
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "Institutsspezifische Risikoanalyse GwG / AML")

    _label_value(pdf, "Dokument-Nr.", "MB-RA-2022-003")
    _label_value(pdf, "Version", "3.0")
    _label_value(pdf, "Erstellt", "Februar 2022")
    _label_value(pdf, "Letzte Überprüfung", "Februar 2022 (vollständige Überarbeitung)")
    _label_value(pdf, "Nächste Überprüfung", "Geplant: Q1 2023 – noch ausstehend")
    _label_value(pdf, "Status", "Genehmigt – Vorstandssitzung 03.02.2022")
    pdf.ln(4)

    _h2(pdf, "1. Einleitung und gesetzliche Grundlage")
    _body(
        pdf,
        "Diese Risikoanalyse wird gemäß § 5 GwG erstellt und dokumentiert die institutsspezifischen "
        "Risiken der Geldwäsche und Terrorismusfinanzierung der Musterbank AG. "
        "Sie bildet die Grundlage für alle weiteren internen Sicherungsmaßnahmen.\n\n"
        "Die Musterbank AG ist ein regional tätiges Kreditinstitut mit Sitz in Frankfurt am Main, "
        "das Retail- und Firmenkundengeschäft betreibt. Das verwaltete Vermögen beläuft sich auf "
        "ca. 2,8 Mrd. EUR. Die Bank unterhält 12 Filialen in Hessen und Rheinland-Pfalz.",
    )

    _h2(pdf, "2. Risikobewertung – Kundensegmente")
    cols = [
        ("Segment", 50),
        ("Anzahl Kunden", 35),
        ("Risikoklasse", 35),
        ("Begründung", 50),
    ]
    _table_header(pdf, cols)
    rows = [
        ("Privatkunden Standard", "27.400", "Niedrig", "Inländisch, kein PEP"),
        ("Privatkunden Vermögend", " 2.100", "Mittel", "Höheres Volumen, Herkunft"),
        ("Firmenkunden KMU", " 7.200", "Mittel", "Strukturkomplexität"),
        ("Firmenkunden Groß", "   340", "Hoch", "Komplexe Strukturen"),
        (
            "PEPs und Familienangehörige",
            "    82",
            "Sehr hoch",
            "§15 GwG verstärkte Pflichten",
        ),
        ("Non-Face-to-Face", " 1.200", "Hoch", "Fehlende persönl. Identif."),
    ]
    widths = [50, 35, 35, 50]
    for i, row in enumerate(rows):
        _table_row(pdf, row, widths, alt=bool(i % 2))
    pdf.ln(4)

    _h2(pdf, "3. Risikobewertung – Produkte und Dienstleistungen")
    _body(
        pdf,
        "Folgende Produkte wurden bewertet:\n"
        "– Girokonto / Zahlungsverkehr: Mittleres Risiko (Volumen, Frequenz)\n"
        "– Sparprodukte / Festgeld: Niedriges Risiko\n"
        "– Konsumentenkredite: Mittleres Risiko (Geldwäsche durch Kreditrückzahlung)\n"
        "– Firmenkredite: Hohes Risiko (Missbrauch für Geldwäsche möglich)\n"
        "– Devisenwechsel: Hohes Risiko (Bargeld, Anonymität)\n"
        "– Online-Banking: Mittleres bis hohes Risiko (Non-Face-to-Face)",
    )

    _h2(pdf, "4. Risikobewertung – Vertriebswege")
    _body(
        pdf,
        "Die Vertriebswege wurden im Rahmen dieser Analyse nur überblicksartig erfasst. "
        "Eine detaillierte Risikobewertung der einzelnen Vertriebskanäle (Filiale, "
        "Online-Banking, Telefonbanking, Maklergeschäft) steht noch aus und soll in "
        "der nächsten Überarbeitung vertieft werden. Vorläufige Einschätzung:\n"
        "– Filialgeschäft: Niedriges bis mittleres Risiko\n"
        "– Online-Banking: Mittleres Risiko (nicht vollständig bewertet)\n"
        "– Drittparteienvertrieb: Nicht bewertet (kein aktiver Kanal)",
    )

    _h2(pdf, "5. Gesamtrisikoeinstufung und Maßnahmen")
    _body(
        pdf,
        "Gesamtrisikoeinstufung der Musterbank AG: MITTEL\n\n"
        "Basierend auf der Analyse wurden folgende Maßnahmen festgelegt:\n"
        "– Verstärkte Sorgfaltspflichten für Kunden der Klassen 'Hoch' und 'Sehr hoch'\n"
        "– Transaktionsmonitoring für alle Kontobewegungen > 10.000 EUR\n"
        "– Jährliche Schulung aller relevanten Mitarbeiter\n"
        "– Quartalsweiser AML-Statusbericht an den Vorstand\n\n"
        "Hinweis zur Aktualität: Die vorliegende Risikoanalyse stammt aus Februar 2022. "
        "Die geplante jährliche Überprüfung gemäß § 5 Abs. 2 GwG wurde bisher nicht durchgeführt. "
        "Eine Aktualisierung war für Q1 2023 geplant, wurde jedoch aufgrund personeller "
        "Engpässe im Compliance-Team verschoben. Die Aktualisierung ist nun für Q2 2025 eingeplant.",
    )
    pdf.output(out)
    log.info("  ✓ %s", out.name)


def create_kyc_handbuch(out: Path):
    """
    S02-01 → KONFORM: Neukunden-Identifizierung vollständig geregelt.
    S02-02 → TEILKONFORM: wB-Ermittlung geregelt, aber Bestandskundenabgleich lückenhaft.
    S02-03 → KONFORM: PEP-Screening-Prozess vorhanden.
    S02-04 → KONFORM: Sanktionslisten-Abgleich automatisiert.
    S02-05 → KONFORM: Hochrisikoländer-Klassifikation vorhanden.
    S02-06 → TEILKONFORM: Re-Identifizierung nur bei Risikoklasse Hoch, nicht alle Kunden.
    """
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "KYC- und Kundensorgfaltspflichten-Handbuch")

    _label_value(pdf, "Dokument-Nr.", "MB-KYC-2024-002")
    _label_value(pdf, "Version", "4.2 (Stand: Januar 2024)")
    _label_value(pdf, "Gilt für", "Alle Mitarbeiter mit Kundenkontakt")
    _label_value(pdf, "Genehmigt", "GwB Dr. Müller, 10.01.2024")
    pdf.ln(4)

    _h2(pdf, "1. Allgemeine Sorgfaltspflichten bei Neukunden (§§ 10-12 GwG)")
    _body(
        pdf,
        "Vor Begründung jeder Geschäftsbeziehung sind folgende Maßnahmen verpflichtend:\n\n"
        "a) Identifizierung natürlicher Personen:\n"
        "– Erhebung von vollständigem Namen, Geburtsdatum, Geburtsort, Staatsangehörigkeit, "
        "Wohnanschrift sowie Art und Nummer des Identitätsnachweises\n"
        "– Gültige Ausweisdokumente: Personalausweis, Reisepass, elektronischer Aufenthaltstitel\n"
        "– Sofortige Verifikation des Dokuments über das IT-System (automatischer OCR-Abgleich)\n\n"
        "b) Identifizierung juristischer Personen:\n"
        "– Erhebung von Firma, Sitz, Registernummer und Rechtsform\n"
        "– Handelsregisterauszug (nicht älter als 3 Monate) oder digitaler Abruf\n"
        "– Identifizierung der gesetzlichen Vertreter (wie natürliche Personen)",
    )

    _h2(pdf, "2. Wirtschaftlich Berechtigte (§ 3 GwG)")
    _body(
        pdf,
        "Bei allen Kunden, die keine natürlichen Personen sind, ist der wirtschaftlich "
        "Berechtigte (wB) zu ermitteln und zu dokumentieren.\n\n"
        "Definition: Als wB gilt jede natürliche Person, die mehr als 25% der Kapitalanteile "
        "oder Stimmrechte hält oder auf sonstige Weise Kontrolle ausübt.\n\n"
        "Prozess bei Neukunden:\n"
        "– Transparenzregister-Abfrage ist für alle juristischen Personen verpflichtend\n"
        "– Abfragenachweis ist in der Kundendokumentation zu hinterlegen\n"
        "– Bei Unstimmigkeiten: Meldung an GwB und ggf. an das Transparenzregister\n\n"
        "WICHTIGER HINWEIS – Bestandskunden:\n"
        "Die systematische Nacherhebung der wB-Dokumentation für Bestandskunden ist noch "
        "nicht abgeschlossen. Zum Stand 31.12.2024 fehlt bei ca. 18% der Firmenkundenkonten "
        "(ca. 1.350 von 7.540 Konten) eine aktuelle Transparenzregister-Abfrage. "
        "Ein Projekt zur Nacherhebung (MB-CDD-BACKFILL-2024) wurde gestartet, "
        "Abschluss geplant für Q3 2025.",
    )

    _h2(pdf, "3. PEP-Screening (§ 15 GwG)")
    _body(
        pdf,
        "Das PEP-Screening ist vollautomatisch in den Kontoeröffnungsprozess integriert:\n\n"
        "– Automatischer Abgleich bei Kontoeröffnung gegen kommerzielle PEP-Datenbank (WorldCheck)\n"
        "– Automatischer Trefferbericht an GwB-Postfach\n"
        "– Täglicher Re-Screening-Lauf gegen aktualisierte PEP-Listen\n"
        "– Bei positivem Treffer: Manuelles Review durch GwB-Team innerhalb 24 Stunden\n"
        "– PEP-Kunden werden automatisch auf Risikoklasse 'Sehr hoch' gesetzt\n"
        "– Verstärkte Sorgfaltspflichten: Herkunftsnachweis Vermögen, Zustimmung GwB zur Kontoeröffnung",
    )

    _h2(pdf, "4. Sanktionslisten-Screening (§ 25h Abs. 2 KWG)")
    _body(
        pdf,
        "Sanktionslisten-Screening wird vollautomatisch durch das Kernbankensystem durchgeführt:\n\n"
        "– Abgeglichene Listen: EU-Finanzsanktionsliste, UN-Sanktionslisten, OFAC SDN, HM Treasury\n"
        "– Echtzeit-Screening bei Transaktionsdurchführung\n"
        "– Tägliche Aktualisierung der Listendaten (automatischer Download)\n"
        "– Matching-Algorithmus mit konfigurierbarem Fuzzy-Score (aktuell: 85% Übereinstimmung)\n"
        "– Bei Treffer: Automatische Transaktionssperrung + sofortige Eskalation an GwB\n"
        "– Dokumentation jedes Treffers im AML-System",
    )

    _h2(pdf, "5. Laufende Überwachung – Bestandskunden (§ 10 Abs. 1 Nr. 4 GwG)")
    _body(
        pdf,
        "Die laufende Überwachung bestehender Geschäftsbeziehungen erfolgt risikobasiert:\n\n"
        "– Kunden Risikoklasse 'Hoch'/'Sehr hoch': Vollständige Re-Identifizierung alle 12 Monate\n"
        "– Kunden Risikoklasse 'Mittel': KYC-Aktualisierung bei wesentlicher Änderung\n"
        "– Kunden Risikoklasse 'Niedrig': Keine proaktive Re-Identifizierung geplant\n\n"
        "Einschränkung: Für Kunden der Risikoklassen 'Niedrig' und 'Mittel' besteht derzeit "
        "kein systematischer Re-Identifizierungsturnus. Dies entspricht nicht vollständig den "
        "Anforderungen des § 10 Abs. 1 Nr. 4 GwG, der eine laufende Überwachung für alle "
        "Kunden vorsieht. Eine Erweiterung des Monitoring-Konzepts ist in Planung.",
    )
    pdf.output(out)
    log.info("  ✓ %s", out.name)


def create_tm_konzept(out: Path):
    """
    S03-01 → KONFORM: Automatisiertes TM-System im Einsatz.
    S03-02 → TEILKONFORM: Szenarien dokumentiert, Schwellenwerte ohne methodische Begründung.
    S03-03 → NICHT_KONFORM: Rückstau von 3 Wochen dokumentiert (Bearbeitungszeit > 7 Tage).
    S03-04 → TEILKONFORM: Validierung 2022 durchgeführt, seitdem keine Wiederholung.
    """
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "Konzept Transaktionsmonitoring (TM-System)")

    _label_value(pdf, "Dokument-Nr.", "MB-TM-2023-001")
    _label_value(pdf, "Version", "2.0 (Stand: März 2023)")
    _label_value(pdf, "System", "FircoSoft TM Suite v7.4")
    _label_value(pdf, "Systemverantwortlicher", "IT-Betrieb / GwB-Team")
    pdf.ln(4)

    _h2(pdf, "1. Systemübersicht")
    _body(
        pdf,
        "Die Musterbank AG betreibt seit 2019 das automatisierte Transaktionsmonitoringsystem "
        "FircoSoft TM Suite v7.4. Das System überwacht alle Kontobewegungen in Echtzeit "
        "und generiert Alerts bei Auffälligkeiten.\n\n"
        "Überwachte Transaktionstypen:\n"
        "– SEPA-Überweisungen (In- und Ausland)\n"
        "– SWIFT-Zahlungsverkehr\n"
        "– Bareinzahlungen und -auszahlungen\n"
        "– Interne Umbuchungen > 5.000 EUR\n"
        "– Wertpapiergeschäfte",
    )

    _h2(pdf, "2. Monitoring-Szenarien")
    _body(pdf, "Das TM-System arbeitet mit 12 aktiven Monitoring-Szenarien:")
    cols = [
        ("Szenario-ID", 25),
        ("Bezeichnung", 70),
        ("Schwellenwert", 40),
        ("Risikoklasse", 35),
    ]
    _table_header(pdf, cols)
    szenarien = [
        ("TM-01", "Grosse Bareinzahlung", "> 9.000 EUR", "Hoch"),
        ("TM-02", "Strukturierung (Smurfing)", "3x > 3.000 in 7T", "Sehr hoch"),
        ("TM-03", "Ungewöhnliche Auslandsüberweisung", "> 25.000 EUR", "Mittel"),
        ("TM-04", "Risikostaat-Transaktion", "> 5.000 EUR", "Hoch"),
        ("TM-05", "Schnelle Kontenentleerung", "> 80% in 24h", "Hoch"),
        ("TM-06", "Neukundenaktivität", "> 50.000 in 30T", "Mittel"),
        ("TM-07", "PEP-Transaktion", "> 10.000 EUR", "Sehr hoch"),
        ("TM-08", "Round-Tripping", "Erkennungsmuster", "Hoch"),
        ("TM-09", "Ungewöhnliche Frequenz", "> 20 TX/Tag", "Mittel"),
        ("TM-10", "Dormant Account Aktivierung", "> 12 Monate inaktiv", "Mittel"),
        ("TM-11", "Correspondent Banking", "> 100.000 EUR", "Hoch"),
        ("TM-12", "Krypto-Börsen-Bezug", "Jede Transaktion", "Sehr hoch"),
    ]
    widths = [25, 70, 40, 35]
    for i, row in enumerate(szenarien):
        _table_row(pdf, row, widths, alt=bool(i % 2))
    pdf.ln(3)
    _body(
        pdf,
        "HINWEIS zur Kalibrierung: Die Schwellenwerte wurden bei Systemeinführung 2019 "
        "festgelegt und seitdem nicht methodisch überprüft. Eine statistische Analyse "
        "(z.B. basierend auf False-Positive-Rate, SAR-Konversionsrate) wurde bisher "
        "nicht durchgeführt. Der aktuelle Schwellenwert TM-01 (9.000 EUR) liegt knapp "
        "unter der gesetzlichen Meldeschwelle von 10.000 EUR und ist nicht explizit begründet.",
    )

    _h2(pdf, "3. Alert-Bearbeitung und aktuelle Situation")
    _body(
        pdf,
        "ALERT-STATISTIK 2024:\n"
        "– Generierte Alerts gesamt:  14.820\n"
        "– Davon geschlossen (FP):    13.490 (91,0%)\n"
        "– Davon eskaliert zu GwB:     1.080 (7,3%)\n"
        "– Davon zu SAR umgewandelt:      89 (0,6%)\n\n"
        "KRITISCHER BEFUND – RÜCKSTAU:\n"
        "Aufgrund von 2 längerfristigen Erkrankungen im Compliance-Team und gestiegener "
        "Alertvolumina (+23% ggü. Vorjahr) besteht seit August 2024 ein signifikanter "
        "Alert-Rückstau. Zum 31.12.2024 waren 847 Alerts älter als 7 Werktage, "
        "davon 312 älter als 21 Werktage. Die durchschnittliche Bearbeitungszeit "
        "betrug im Q4 2024: 19 Tage. Gemäß interner Richtlinie sind Alerts innerhalb "
        "von 5 Werktagen zu schließen. Eine Abarbeitung des Rückstaus wird bis Q2 2025 angestrebt.",
    )

    _h2(pdf, "4. Systemvalidierung")
    _body(
        pdf,
        "Eine Validierung der TM-System-Wirksamkeit wurde zuletzt im Oktober 2022 "
        "durch die Interne Revision durchgeführt. Der Bericht (IR-2022-TM-01) kam zu "
        "dem Ergebnis, dass das System grundsätzlich funktionsfähig ist, jedoch "
        "Optimierungspotenzial bei den Szenarien TM-03 und TM-09 besteht.\n\n"
        "Die empfohlenen Anpassungen wurden teilweise umgesetzt. Eine vollständige "
        "Folge-Validierung, die gemäß Prüfungsplan für 2024 vorgesehen war, wurde "
        "aufgrund von Kapazitätsengpässen auf 2025 verschoben.",
    )
    pdf.output(out)
    log.info("  ✓ %s", out.name)


def create_sar_verfahren(out: Path):
    """
    S05-01 → KONFORM: Dokumentierter SAR-Prozess vorhanden.
    S05-02 → KONFORM: Fristen eingehalten (durch SAR-Statistik belegt).
    S05-03 → KONFORM: Tipping-Off-Verbot explizit geregelt.
    """
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "Verfahrensanweisung Verdachtsmeldewesen (§ 43 GwG)")

    _label_value(pdf, "Dokument-Nr.", "MB-SAR-2024-001")
    _label_value(pdf, "Version", "3.1 (Stand: April 2024)")
    _label_value(pdf, "Genehmigt", "GwB Dr. Müller und Vorstand, 15.04.2024")
    _label_value(
        pdf, "Gilt für", "GwB-Team, Compliance, alle Mitarbeiter mit Meldeverdacht"
    )
    pdf.ln(4)

    _h2(pdf, "1. Meldeverpflichtung")
    _body(
        pdf,
        "Die Musterbank AG ist gemäß § 43 GwG verpflichtet, bei Verdacht auf Geldwäsche "
        "oder Terrorismusfinanzierung unverzüglich eine Verdachtsmeldung an die Financial "
        "Intelligence Unit (FIU) beim Zoll zu erstatten.\n\n"
        "Die Meldepflicht gilt für alle Mitarbeiter. Verdachtsmomente sind unverzüglich "
        "dem GwB zu melden. Die Entscheidung über die Erstattung einer SAR trifft ausschließlich der GwB.",
    )

    _h2(pdf, "2. Meldeprozess (Schritt für Schritt)")
    _body(
        pdf,
        "Schritt 1: Mitarbeiter identifiziert Verdachtsmoment\n"
        "           → Interne Meldung an GwB via verschlüsseltes E-Mail-Postfach (gwb@musterbank.de)\n"
        "           → Alternativ: Direktanruf GwB-Hotline (intern 9099)\n\n"
        "Schritt 2: GwB-Team nimmt Sachverhalt auf (innerhalb 2 Werktage)\n"
        "           → Erste Einschätzung: Verdacht begründet ja/nein?\n"
        "           → Aktenanlage im AML-System (Referenznummer wird vergeben)\n\n"
        "Schritt 3: GwB entscheidet über Meldung (innerhalb weiterer 2 Werktage)\n"
        "           → Bei Entscheidung zur Meldung: Erstellung SAR-Formular in goAML\n"
        "           → Einreichung via goAML-Portal der FIU\n"
        "           → Dokumentation des Eingangsbestätigungs-Timestamps\n\n"
        "Schritt 4: Post-Meldung\n"
        "           → Dokumentation im AML-System\n"
        "           → Information an Vorstand (monatlicher Bericht)\n"
        "           → Kontoführung gemäß § 46 GwG (Einfrierung wenn angeordnet)",
    )

    _h2(pdf, "3. Tipping-Off-Verbot (§ 47 GwG)")
    _body(
        pdf,
        "Es ist STRENG VERBOTEN, den Kunden oder Dritte über eine erstattete oder "
        "beabsichtigte Verdachtsmeldung zu informieren (§ 47 Abs. 1 GwG). "
        "Verstöße sind strafbewehrt (§ 17 GwG) und können persönliche Haftung auslösen.\n\n"
        "Schulung: Das Tipping-Off-Verbot ist Bestandteil jeder jährlichen GwG-Schulung "
        "und wird in der Einarbeitung neuer Mitarbeiter explizit behandelt.\n\n"
        "Im Fall einer Verdachtsmeldung wird der betroffene Mitarbeiter durch den GwB "
        "persönlich über das Tipping-Off-Verbot belehrt und eine Schweigepflichts-Bestätigung "
        "eingeholt (Formular MB-SAR-F01).",
    )

    _h2(pdf, "4. Statistik 2024")
    _body(
        pdf,
        "Verdachtsmeldungen 2024:\n"
        "– Interne Meldungen eingegangen:    156\n"
        "– Davon SAR erstattet:               89 (57,1%)\n"
        "– Davon nicht erstattet (kein Verd.): 67 (42,9%)\n"
        "– Durchschnittliche Bearbeitungszeit:  3,2 Werktage (Ziel: < 5)\n"
        "– Einreichung innerhalb Frist (§ 43): 100% der erstatteten SARs\n"
        "– Rückmeldungen der FIU erhalten:     12 (13,5%)",
    )
    pdf.output(out)
    log.info("  ✓ %s", out.name)


def create_schulungskonzept(out: Path):
    """
    S06-01 → TEILKONFORM: Schulungen finden statt, aber nur 72% Abdeckung 2024.
    S06-02 → KONFORM: Schulungsinhalte aktuell, neue Themen abgedeckt.
    """
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "Schulungskonzept AML / GwG")

    _label_value(pdf, "Dokument-Nr.", "MB-SCH-2024-001")
    _label_value(pdf, "Version", "2.3 (Stand: Januar 2024)")
    _label_value(
        pdf,
        "Zielgruppen",
        "Alle Mitarbeiter (Pflichtschulung); vertieft für Vertrieb, Compliance, IT",
    )
    _label_value(pdf, "Genehmigt", "GwB Dr. Müller und HR, 08.01.2024")
    pdf.ln(4)

    _h2(pdf, "1. Schulungspflicht und Zielgruppen")
    _body(
        pdf,
        "Gemäß § 6 Abs. 2 Nr. 6 GwG sind alle Mitarbeiter, die mit der Durchführung von "
        "Transaktionen oder der Begründung von Geschäftsbeziehungen befasst sind, "
        "regelmäßig zu schulen. Die Musterbank AG hat folgende Schulungsstruktur eingeführt:\n\n"
        "Pflichtkurs (alle Mitarbeiter, jährlich, 45 Min.):\n"
        "– GwG-Grundlagen, Typologien, Meldeverpflichtung, Tipping-Off\n\n"
        "Aufbaukurs Vertrieb (Kundenberater, jährlich, 90 Min.):\n"
        "– KYC, PEP-Erkennung, verstärkte Sorgfaltspflichten\n\n"
        "Spezialkurs Compliance/GwB-Team (halbjährlich, 3h):\n"
        "– Aktuelle BaFin-Auslegungs- und Anwendungshinweise, SAR-Praxis, TM",
    )

    _h2(pdf, "2. Schulungsinhalte 2024 (Pflichtkurs)")
    _body(
        pdf,
        "Der Pflichtkurs 2024 deckt folgende aktuelle Themen ab:\n"
        "– Aktualisierungen AMLA (EU-Geldwäschebehörde) und 6. EU-AML-Richtlinie\n"
        "– Neue BaFin-Auslegungs- und Anwendungshinweise (Stand: 2023)\n"
        "– Kryptowerte als Geldwäscherisiko: Fallbeispiele\n"
        "– Aktuelle Geldwäsche-Typologien (FATF Mutual Evaluation Germany 2022)\n"
        "– Fälle aus der Bankpraxis: 3 anonymisierte Fallstudien aus 2023\n"
        "– Prüfungsrecht der BaFin, Sonderprüfungen nach § 44 KWG\n\n"
        "Externe Schulungsanbieter: ACAMS Online-Kurs (Pflichtkurs), ICA für Spezialkurse",
    )

    _h2(pdf, "3. Schulungsdurchführung und -nachweise")
    _body(
        pdf,
        "Schulungen werden über das LMS (Learning Management System, SAP SuccessFactors) "
        "organisiert und dokumentiert. Jeder abgeschlossene Kurs wird automatisch "
        "mit Zeitstempel, Testergebnis und Teilnehmerdaten archiviert.\n\n"
        "SCHULUNGSERGEBNISSE 2024 (Stichtag 31.12.2024):\n"
        "Pflichtkurs GwG Grundlagen:\n"
        "– Pflichtig: 520 Mitarbeiter\n"
        "– Abgeschlossen: 374 Mitarbeiter (72,0%)\n"
        "– Nicht abgeschlossen: 146 Mitarbeiter (28,0%) – davon:\n"
        "    * 41 Langzeitkranke (> 6 Wochen)\n"
        "    * 38 im Berichtsjahr neu eingestellt (Onboarding-Schulung Q1 2025 geplant)\n"
        "    * 67 ohne dokumentierten Abschlussgrund\n\n"
        "Aufbaukurs Vertrieb:\n"
        "– Pflichtig: 180 Mitarbeiter\n"
        "– Abgeschlossen: 164 Mitarbeiter (91,1%)\n\n"
        "Maßnahme: GwB hat am 15.01.2025 eine Erinnerungsmail an die 67 Mitarbeiter ohne "
        "Abschlussgrund versandt. Frist: 28.02.2025.",
    )
    pdf.output(out)
    log.info("  ✓ %s", out.name)


def create_aufbewahrung_iks(out: Path):
    """
    S07-01 → KONFORM: Aufbewahrungspflichten formal dokumentiert.
    S07-02 → KONFORM: Bereitstellungsprozess für BaFin-Anfragen vorhanden.
    S08-01 → NICHT_PRÜFBAR: IR-Prüfbericht für AML nicht vorgelegt (nur Verweis).
    S08-02 → NICHT_KONFORM: Kein Maßnahmentracking, offene Punkte nicht systematisch verfolgt.
    """
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "Aufbewahrungspflichten und Interne Revision AML")

    _label_value(pdf, "Dokument-Nr.", "MB-IKS-2023-004")
    _label_value(pdf, "Version", "1.2 (Stand: Juni 2023)")
    _label_value(pdf, "Thema", "Aufbewahrung (§8 GwG) und Governance AML-Framework")
    pdf.ln(4)

    _h2(pdf, "1. Aufbewahrungspflichten (§ 8 GwG)")
    _body(
        pdf,
        "Folgende Unterlagen werden revisionssicher für mindestens 5 Jahre aufbewahrt:\n"
        "– Identifizierungsunterlagen (Ausweiskopien, Handelsregisterauszüge)\n"
        "– Nachweise über wirtschaftlich Berechtigte\n"
        "– Transaktionsdaten für alle überwachungspflichtigen Geschäfte\n"
        "– SAR-Dokumentation (Meldung, interne Einschätzung, Kommunikation FIU)\n"
        "– TM-Alert-Dokumentation (Alerts, Bearbeitungsnotizen, Schliessentscheidung)\n"
        "– Schulungsnachweise (LMS-Export)\n\n"
        "Aufbewahrungssystem: IBM FileNet P8 (revisionssicheres Dokumentenmanagementsystem)\n"
        "Löschfristen: Automatisierte Löschung nach Ablauf der gesetzlichen Mindestfrist + "
        "1 Jahr Puffer. Löschprotokoll wird jährlich vom GwB geprüft.",
    )

    _h2(pdf, "2. BaFin-Bereitstellung")
    _body(
        pdf,
        "Im Falle einer BaFin-Anfrage oder Sonderprüfung (§ 44 KWG) ist sichergestellt:\n\n"
        "– Innerhalb 24h: Bereitstellung aller angeforderten Unterlagen in elektronischer Form\n"
        "– Ansprechpartner: GwB Dr. Müller (Primär), Stellvertreterin S. Becker (Sekundär)\n"
        "– Dedizierter Prüfungsraum: Konferenzraum 3B, Hauptstelle Frankfurt\n"
        "– Datenverschlüsselung für elektronische Übermittlung: PGP/GPG-Standard\n\n"
        "Referenz: Letzte BaFin-Anfrage (nicht öffentlich, AZ 2023/MBK/0012) wurde in "
        "18 Stunden vollständig beantwortet.",
    )

    _h2(pdf, "3. Interne Revision AML-Framework")
    _body(
        pdf,
        "Prüfplanung: Gemäß Prüfungsplan 2024 war eine Prüfung des AML-Frameworks durch "
        "die Interne Revision für Q3 2024 vorgesehen. Diese Prüfung wurde aufgrund anderer "
        "aufsichtsrechtlich priorisierter Prüfungsvorhaben (MaRisk-Prüfung, DORA-Readiness) "
        "auf Q1 2025 verschoben.\n\n"
        "Letzte AML-Prüfung durch IR: Oktober 2022 (IR-Bericht 2022-AML-01).\n"
        "Der IR-Bericht 2022 liegt in der Revisionsabteilung vor, wurde aber für diese "
        "Prüfung nicht als separates Dokument übergeben.\n\n"
        "MASSNAHMENTRACKING:\n"
        "Aus der IR-Prüfung 2022 wurden 7 Maßnahmen vereinbart. Ein systematisches, "
        "IT-gestütztes Tracking dieser Maßnahmen existiert nicht. Die Nachverfolgung "
        "erfolgt über eine Excel-Datei (IR-Massnahmen.xlsx), die zuletzt im März 2023 "
        "aktualisiert wurde. Zum aktuellen Stand der offenen Punkte liegen keine "
        "dokumentierten Nachweise vor.",
    )
    pdf.output(out)
    log.info("  ✓ %s", out.name)


def create_gwb_jahresbericht(out: Path):
    """Stützt S04-03 (KONFORM) und liefert Kontext für S03, S05, S06."""
    pdf = _pdf_new()
    pdf.add_page()
    _h1(pdf, "GwB-Jahresbericht 2024 – Zusammenfassung Vorstand")

    _label_value(pdf, "Dokument-Nr.", "MB-GWB-JAB-2024")
    _label_value(pdf, "Autor", "Dr. Thomas Müller, Geldwäschebeauftragter")
    _label_value(pdf, "Datum", "28. Januar 2025")
    _label_value(
        pdf, "Genehmigt", "Gesamtvorstand, 28.01.2025 (TOP 3 Vorstandssitzung)"
    )
    pdf.ln(4)

    _h2(pdf, "1. Zusammenfassung")
    _body(
        pdf,
        "Das Jahr 2024 war geprägt von steigendem Transaktionsvolumen (+14%), "
        "wachsenden regulatorischen Anforderungen (AMLA-Trilog, DORA) und "
        "personellen Herausforderungen im GwB-Team. "
        "Trotz dieser Rahmenbedingungen wurden alle gesetzlichen Kernpflichten erfüllt. "
        "Handlungsbedarf besteht insbesondere bei:\n"
        "– Alert-Rückstau im TM-System\n"
        "– Vollständige wB-Dokumentation bei Bestandskunden\n"
        "– Ressourcenausstattung des GwB-Teams\n"
        "– Aktualisierung der Risikoanalyse (überfällig)",
    )

    _h2(pdf, "2. Kennzahlen 2024")
    cols = [("Kennzahl", 90), ("Wert 2024", 40), ("Wert 2023", 40)]
    _table_header(pdf, cols)
    kpis = [
        ("Neue Geschäftsbeziehungen", "4.820", "4.210"),
        ("KYC-Vollständigkeit Neukunden", "100%", "100%"),
        ("wB-Dokumentation Bestandskunden", "82%", "79%"),
        ("TM-Alerts generiert", "14.820", "12.030"),
        ("Alert-False-Positive-Rate", "91,0%", "89,3%"),
        ("Durchschn. Alert-Bearbeitungszeit", "19 Tage*", "6 Tage"),
        ("Interne Meldungen (Verdacht)", "156", "121"),
        ("Erstattete SARs", "89", "74"),
        ("SAR-Fristerfüllung", "100%", "100%"),
        ("GwG-Schulungsquote (Pflichtkurs)", "72%", "94%"),
        ("PEP-Kunden (Stand 31.12.)", "82", "79"),
        ("Sanktionslisten-Treffer", "3 (FP)", "1 (FP)"),
    ]
    widths = [90, 40, 40]
    for i, row in enumerate(kpis):
        _table_row(pdf, row, widths, alt=bool(i % 2))
    pdf.ln(2)
    _body(
        pdf,
        "* Anstieg durch krankheitsbedingte Personalausfälle im Compliance-Team (Q3-Q4 2024).",
    )

    _h2(pdf, "3. Wesentliche Ereignisse und Maßnahmen")
    _body(
        pdf,
        "März 2024: BaFin-Informationsveranstaltung zu AMLA und 6. AML-Paket – teilgenommen\n"
        "April 2024: Aktualisierung Schulungsmaterial und SAR-Verfahrensanweisung\n"
        "Juli 2024: Krankenstand GwB-Team: 1 von 2 Mitarbeitern längerfristig ausgefallen\n"
        "August 2024: Beginn Alert-Rückstau – Gegenmaßnahmen eingeleitet (Überst., Priorisierung)\n"
        "Oktober 2024: Ersatz für erkrankten Mitarbeiter: Zeitarbeitskraft (ohne GwG-Zertifizierung)\n"
        "November 2024: WorldCheck-Datenbankupdate – 3 neue PEP-Treffer identifiziert und bearbeitet\n"
        "Dezember 2024: Projekt MB-CDD-BACKFILL-2024 gestartet (wB-Nacherhebung Bestandskunden)",
    )

    _h2(pdf, "4. Empfehlungen an den Vorstand")
    _body(
        pdf,
        "1. RESSOURCEN (dringend): Aufstockung GwB-Team um mindestens 1,5 FTE\n"
        "   Begründung: 520 MA und 38.000 Kunden sind mit 2 Personen nicht angemessen zu betreuen\n\n"
        "2. RISIKOANALYSE: Sofortiger Start der Aktualisierung, Abschluss bis Q2 2025\n"
        "   Begründung: Letzte vollständige Überarbeitung Feb 2022 – nicht § 5 GwG-konform\n\n"
        "3. TM-RÜCKSTAU: Temporäre externe Unterstützung bis Alert-Rückstand abgebaut\n"
        "   Begründung: 847 Alerts > 7 Werktage stellen Compliance-Risiko dar\n\n"
        "4. MASSNAHMENTRACKING: Einführung eines IT-gestützten Systems\n"
        "   Begründung: Excel-Lösung nicht revisionssicher und nicht vollständig aktuell",
    )
    pdf.output(out)
    log.info("  ✓ %s", out.name)


# ------------------------------------------------------------------ #
# Excel / CSV
# ------------------------------------------------------------------ #


def create_tm_alerts_excel(out: Path):
    """TM-Alert-Statistik 2024 – stützt S03-03 (NICHT_KONFORM) und S03-02."""
    rng = random.Random(42)

    # Monatliche Alert-Daten
    months = [f"{m:02d}/2024" for m in range(1, 13)]
    generated = [1050, 1080, 1150, 1180, 1210, 1250, 1310, 1380, 1420, 1410, 1420, 960]
    closed_fp = [int(g * rng.uniform(0.87, 0.93)) for g in generated]
    escalated = [g - fp - rng.randint(5, 20) for g, fp in zip(generated, closed_fp)]
    to_sar = [max(4, int(e * rng.uniform(0.07, 0.11))) for e in escalated]
    avg_days = [4.2, 4.8, 5.1, 5.3, 5.0, 5.4, 5.8, 14.2, 18.7, 21.3, 23.1, 19.4]
    backlog_over7 = [0, 0, 0, 0, 0, 0, 12, 89, 234, 512, 698, 847]

    df_monthly = pd.DataFrame(
        {
            "Monat": months,
            "Alerts_generiert": generated,
            "Alerts_FP_geschlossen": closed_fp,
            "Alerts_eskaliert_GwB": escalated,
            "Daraus_SAR": to_sar,
            "Durchschn_Bearbeitungszeit_Tage": avg_days,
            "Rückstau_über_7_Werktage": backlog_over7,
        }
    )

    # Szenario-Auswertung
    szenarien_ids = [f"TM-{i:02d}" for i in range(1, 13)]
    szenarien_names = [
        "Grosse Bareinzahlung",
        "Strukturierung (Smurfing)",
        "Auslandsüberweisung",
        "Risikostaat-Transaktion",
        "Schnelle Kontenentleerung",
        "Neukundenaktivität",
        "PEP-Transaktion",
        "Round-Tripping",
        "Ungewöhnliche Frequenz",
        "Dormant Account",
        "Correspondent Banking",
        "Krypto-Bezug",
    ]
    total_alerts_szenario = [rng.randint(400, 2800) for _ in range(12)]
    fp_rate = [rng.uniform(0.82, 0.97) for _ in range(12)]
    sar_count = [rng.randint(2, 18) for _ in range(12)]

    df_szenarien = pd.DataFrame(
        {
            "Szenario_ID": szenarien_ids,
            "Bezeichnung": szenarien_names,
            "Alerts_2024": total_alerts_szenario,
            "FP_Rate_Prozent": [round(r * 100, 1) for r in fp_rate],
            "SAR_Meldungen": sar_count,
            "Konversionsrate_SAR_Prozent": [
                round(s / a * 100, 2) for s, a in zip(sar_count, total_alerts_szenario)
            ],
        }
    )

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_monthly.to_excel(writer, sheet_name="Monatliche Übersicht", index=False)
        df_szenarien.to_excel(writer, sheet_name="Szenario-Auswertung", index=False)
    log.info("  ✓ %s", out.name)


def create_schulungsmatrix_excel(out: Path):
    """Schulungsmatrix 2024 – stützt S06-01 (TEILKONFORM: 72%)."""
    rng = random.Random(7)

    abteilungen = [
        ("Privatkundenberatung", 120, "Pflicht + Aufbau"),
        ("Firmenkundenberatung", 60, "Pflicht + Aufbau"),
        ("Marktfolge Kredit", 45, "Pflicht"),
        ("Zahlungsverkehr", 38, "Pflicht"),
        ("IT und Systeme", 52, "Pflicht"),
        ("Personal und Recht", 28, "Pflicht"),
        ("Interne Revision", 18, "Pflicht + Spezialkurs"),
        ("Compliance / GwB-Team", 4, "Pflicht + Aufbau + Spezialkurs"),
        ("Vorstand und Stäbe", 12, "Pflicht"),
        ("Filialen (Kassenpersonal)", 143, "Pflicht"),
    ]

    rows = []
    for abt, count, kurs_typ in abteilungen:
        base_quote = rng.uniform(0.65, 0.98)
        abgeschlossen = int(count * base_quote)
        ausstehend = count - abgeschlossen
        krank = min(ausstehend, int(rng.uniform(0.2, 0.4) * ausstehend))
        neu = min(ausstehend - krank, int(rng.uniform(0.15, 0.35) * ausstehend))
        unbekannt = ausstehend - krank - neu
        rows.append(
            {
                "Abteilung": abt,
                "Kurstyp": kurs_typ,
                "Mitarbeiter_gesamt": count,
                "Abgeschlossen": abgeschlossen,
                "Ausstehend_gesamt": ausstehend,
                "Ausstehend_Krankheit": krank,
                "Ausstehend_Neuzugang": neu,
                "Ausstehend_ohne_Grund": unbekannt,
                "Quote_Prozent": round(abgeschlossen / count * 100, 1),
            }
        )

    df = pd.DataFrame(rows)

    # Gesamtzeile
    total = df[
        [
            "Mitarbeiter_gesamt",
            "Abgeschlossen",
            "Ausstehend_gesamt",
            "Ausstehend_Krankheit",
            "Ausstehend_Neuzugang",
            "Ausstehend_ohne_Grund",
        ]
    ].sum()
    total_row = {
        "Abteilung": "GESAMT",
        "Kurstyp": "",
        "Mitarbeiter_gesamt": total["Mitarbeiter_gesamt"],
        "Abgeschlossen": total["Abgeschlossen"],
        "Ausstehend_gesamt": total["Ausstehend_gesamt"],
        "Ausstehend_Krankheit": total["Ausstehend_Krankheit"],
        "Ausstehend_Neuzugang": total["Ausstehend_Neuzugang"],
        "Ausstehend_ohne_Grund": total["Ausstehend_ohne_Grund"],
        "Quote_Prozent": round(
            total["Abgeschlossen"] / total["Mitarbeiter_gesamt"] * 100, 1
        ),
    }
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Schulungsmatrix 2024", index=False)
    log.info("  ✓ %s", out.name)


def create_sar_statistik_csv(out: Path):
    """SAR-Statistik 2024 – stützt S05-02 (KONFORM: Fristen eingehalten)."""
    rng = random.Random(99)
    rows = []
    sar_id = 1000
    base_date = date(2024, 1, 3)
    for _ in range(89):
        days_offset = rng.randint(0, 358)
        internal_date = base_date + timedelta(days=days_offset)
        bearbeitungszeit = rng.randint(1, 5)  # immer < 5 Tage → KONFORM
        meldedatum = internal_date + timedelta(days=bearbeitungszeit)
        szenario = rng.choice(
            [
                "Strukturierung",
                "Bareinzahlung > 9.000 EUR",
                "Auslandsüberweisung",
                "PEP-Transaktion",
                "Kontoverhalten auffällig",
                "Mitarbeitermeldung",
            ]
        )
        rows.append(
            {
                "SAR_ID": f"SAR-2024-{sar_id}",
                "Datum_interne_Meldung": internal_date.isoformat(),
                "Datum_FIU_Meldung": meldedatum.isoformat(),
                "Bearbeitungszeit_Tage": bearbeitungszeit,
                "Auslösendes_Szenario": szenario,
                "goAML_Bestätigung": f"goAML-{rng.randint(100000, 999999)}",
                "Frist_eingehalten": "Ja",
            }
        )
        sar_id += 1

    df = pd.DataFrame(rows)
    df.to_csv(out, index=False, encoding="utf-8")
    log.info("  ✓ %s", out.name)


# ------------------------------------------------------------------ #
# Interview-Dokumente (JSON)
# ------------------------------------------------------------------ #


def create_interview_gwb(out: Path):
    """Interview mit GwB Dr. Müller – stützt S04-02, S01-02, S02-02, S08."""
    data = {
        "meta": {
            "dokument_typ": "Prüfungsinterview",
            "institut": "Musterbank AG",
            "datum": "2025-02-12",
            "uhrzeit": "10:00 – 12:30 Uhr",
            "ort": "Hauptstelle Frankfurt, Konferenzraum 3B",
            "interviewer": "BaFin-Prüfungsteam (anonym)",
            "interviewter": "Dr. Thomas Müller, Geldwäschebeauftragter",
            "beisitzer": "Sandra Becker, stellv. GwB",
            "protokollführer": "Compliance Assistenz",
        },
        "fragen_antworten": [
            {
                "id": "I-01",
                "prueffeld_referenz": "S04-01",
                "frage": "Seit wann sind Sie als GwB der Musterbank AG bestellt und wann erfolgte die BaFin-Meldung?",
                "antwort": "Ich bin seit dem 15. März 2022 als GwB bestellt. Die BaFin-Meldung erfolgte am 16. März 2022 und wurde unter dem Aktenzeichen 2022/GWB/0471 bestätigt. Vor mir war Frau Klaudia Brandt GwB, die in Rente gegangen ist. Die Übergabe fand über 4 Wochen statt.",
                "kommentar": "Bestellungsurkunde liegt vor. BaFin-Bestätigung wurde vorgelegt.",
            },
            {
                "id": "I-02",
                "prueffeld_referenz": "S04-02",
                "frage": "Wie viele Mitarbeiter stehen Ihnen im GwB-Team zur Verfügung? Empfinden Sie die Ressourcenausstattung als ausreichend?",
                "antwort": "Derzeit arbeiten wir zu zweit: Ich und Frau Becker als Stellvertreterin. Das ist für 520 Mitarbeiter und 38.000 Kunden, ehrlich gesagt, sehr eng. In der zweiten Jahreshälfte 2024 hatte Frau Becker noch einen längeren Krankenstand, das hat uns erheblich unter Druck gesetzt. Ich habe den Vorstand mehrfach auf die Personalsituation hingewiesen, zuletzt im Jahresbericht 2024. Eine Aufstockung ist geplant, aber noch nicht konkretisiert.",
                "kommentar": "Ressourcenproblem bestätigt. Kein formeller Stellenplan für GwB-Bereich vorgefunden.",
            },
            {
                "id": "I-03",
                "prueffeld_referenz": "S01-02",
                "frage": "Wann wurde die institutsspezifische Risikoanalyse zuletzt vollständig überarbeitet?",
                "antwort": "Die aktuelle Version stammt aus Februar 2022. Ich weiß, das ist zu lang. Wir hatten für 2023 eine Überarbeitung geplant, aber das wurde immer wieder geschoben. Jetzt haben wir Q2 2025 als festes Datum gesetzt. Ich habe dafür auch externe Unterstützung beantragt.",
                "kommentar": "Risikoanalyse von Feb 2022 liegt vor. Keine Aktualisierung in 36 Monaten. Kritischer Befund.",
            },
            {
                "id": "I-04",
                "prueffeld_referenz": "S02-02",
                "frage": "Wie stellt die Bank sicher, dass wirtschaftlich Berechtigte bei allen Kunden vollständig dokumentiert sind?",
                "antwort": "Bei Neukunden läuft das hervorragend, da ist der Prozess ins Kernbankensystem integriert und kann gar nicht übersprungen werden. Das Problem sind die Bestandskunden, die vor 2020 eingerichtet wurden. Da fehlt bei etwa 18% der Firmenkonten eine aktuelle Transparenzregister-Abfrage. Wir haben Ende 2024 das Projekt BACKFILL gestartet, das soll bis Q3 2025 abgeschlossen sein.",
                "kommentar": "wB-Lücke bei Bestandskunden bestätigt (ca. 1.350 Konten). Projekt gestartet, noch nicht abgeschlossen.",
            },
            {
                "id": "I-05",
                "prueffeld_referenz": "S03-03",
                "frage": "Wie hat sich die Alert-Bearbeitungszeit im Laufe des Jahres 2024 entwickelt?",
                "antwort": "Das war unser größtes Problem 2024. Bis Juli lagen wir noch im Ziel – unter fünf Tagen. Dann ist Frau Becker ausgefallen und gleichzeitig ist das Alertvolumen gestiegen, weil wir neue TM-Szenarien aktiviert haben. Im Q4 lagen wir teilweise bei über drei Wochen Bearbeitungszeit. Das ist natürlich inakzeptabel. Wir haben im Oktober eine externe Zeitarbeitskraft eingesetzt, aber die war leider nicht einschlägig vorgebildet und musste erst eingearbeitet werden.",
                "kommentar": "Rückstau mit Bearbeitungszeiten von bis zu 23 Tagen bestätigt. TM-Konzept dokumentiert ebenfalls.",
            },
            {
                "id": "I-06",
                "prueffeld_referenz": "S08-01",
                "frage": "Wann hat die Interne Revision das AML-Framework zuletzt geprüft?",
                "antwort": "Die letzte vollständige AML-Prüfung war Oktober 2022. Die Prüfung für 2024 wurde auf 2025 verschoben – das war auch eine Kapazitätsentscheidung der Revision. Der Bericht 2022 liegt bei der Revision. Den kann ich Ihnen inhaltlich nicht vollständig referieren, weil ich ihn nicht in meinen Unterlagen habe.",
                "kommentar": "IR-Bericht 2022 nicht als separates Dokument vorgefunden. Prüfungslücke von über 2 Jahren.",
            },
            {
                "id": "I-07",
                "prueffeld_referenz": "S08-02",
                "frage": "Wie werden Maßnahmen aus Prüfungen systematisch nachverfolgt?",
                "antwort": "Wir haben dafür eine Excel-Tabelle, die von der Compliance geführt wird. Die ist leider seit März 2023 nicht mehr vollständig aktualisiert worden. Ich hatte das als einen der Punkte für das Q1 2025 auf dem Plan. Ein richtiges Workflow-Tool haben wir dafür nicht.",
                "kommentar": "Kein systematisches Maßnahmentracking. Excel-Liste veraltet. Kritischer Befund.",
            },
        ],
    }
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info("  ✓ %s", out.name)


def create_interview_compliance(out: Path):
    """Interview mit Compliance Officer – stützt S06, S05, S02."""
    data = {
        "meta": {
            "dokument_typ": "Prüfungsinterview",
            "institut": "Musterbank AG",
            "datum": "2025-02-12",
            "uhrzeit": "14:00 – 15:30 Uhr",
            "ort": "Hauptstelle Frankfurt, Konferenzraum 3B",
            "interviewer": "BaFin-Prüfungsteam (anonym)",
            "interviewter": "Sandra Becker, stellv. GwB und Compliance Officer",
        },
        "fragen_antworten": [
            {
                "id": "C-01",
                "prueffeld_referenz": "S06-01",
                "frage": "Wie hoch war die Schulungsquote für den GwG-Pflichtkurs 2024, und was sind die Gründe für Nicht-Teilnahmen?",
                "antwort": "Wir sind auf 72% gekommen. Das ist schlechter als 2023, wo wir 94% hatten. Die Hauptgründe sind Langzeitkranke, Neuzugänge und leider auch eine Gruppe von 67 Mitarbeitern, für die wir keinen dokumentierten Grund haben. Ich habe im Januar 2025 persönlich nachgefasst und eine Frist bis Ende Februar gesetzt.",
                "kommentar": "Schulungsquote 72% bestätigt. LMS-Export liegt vor.",
            },
            {
                "id": "C-02",
                "prueffeld_referenz": "S06-02",
                "frage": "Sind die Schulungsinhalte aktuell? Wurden neue regulatorische Entwicklungen berücksichtigt?",
                "antwort": "Ja, das würde ich mit gutem Gewissen sagen. Wir haben den Kurs im April 2024 aktualisiert – AMLA, 6. AML-Paket, Kryptotypologien sind alle drin. Wir kaufen den Grundkurs bei ACAMS ein, die halten das automatisch aktuell.",
                "kommentar": "Schulungsunterlagen wurden gesichtet. Inhalte aktuell.",
            },
            {
                "id": "C-03",
                "prueffeld_referenz": "S05-01",
                "frage": "Beschreiben Sie den internen Prozess, wenn ein Mitarbeiter einen Geldwäscheverdacht meldet.",
                "antwort": "Der Mitarbeiter meldet entweder per verschlüsselter E-Mail an gwb@musterbank.de oder ruft direkt die GwB-Hotline an. Das GwB-Team nimmt den Sachverhalt innerhalb von zwei Werktagen auf, bewertet, ob der Verdacht hinreichend begründet ist, und entscheidet dann über die SAR-Erstellung in goAML. Wir haben eine Verfahrensanweisung dazu – MB-SAR-2024-001.",
                "kommentar": "Prozess gut dokumentiert und beschrieben. SAR-Verfahrensanweisung liegt vor.",
            },
            {
                "id": "C-04",
                "prueffeld_referenz": "S02-04",
                "frage": "Wie funktioniert das Sanktionslisten-Screening in der Praxis?",
                "antwort": "Das ist vollautomatisiert. Jede Transaktion wird in Echtzeit gegen die EU, UN und OFAC Listen gescreent. Die Listen werden täglich aktualisiert. Bei einem Treffer wird die Transaktion sofort gesperrt und ich sowie Dr. Müller bekommen eine automatische Benachrichtigung. 2024 hatten wir drei Treffer – alle als False Positive bestätigt, alle dokumentiert.",
                "kommentar": "Sanktionsscreening automatisiert bestätigt. Drei FP-Treffer dokumentiert.",
            },
        ],
    }
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info("  ✓ %s", out.name)


# ------------------------------------------------------------------ #
# Log-Dateien
# ------------------------------------------------------------------ #


def create_tm_system_log(out: Path):
    """TM-System-Auditlog – stützt S03-01 (System läuft)."""
    rng = random.Random(55)
    lines = [
        "# FircoSoft TM Suite v7.4 – System Audit Log",
        "# Musterbank AG | Exportiert: 2025-01-05 06:00:02",
        "# Zeitraum: 01.12.2024 – 31.12.2024",
        "",
    ]
    event_types = [
        ("INFO", "System Health Check: All 12 TM-Szenarien aktiv"),
        ("INFO", "Datenbankverbindung: OK (response 3ms)"),
        ("INFO", "Regelwerk-Update: Keine Änderungen"),
        ("WARN", "Alert-Queue-Tiefe: {depth} Einträge (Warnschwelle: 500)"),
        ("INFO", "Batch-Verarbeitung abgeschlossen: {count} Transaktionen geprüft"),
        ("INFO", "Sanktionslisten-Update erfolgreich importiert"),
        (
            "ALERT",
            "Neuer Alert generiert: {id} | Szenario: {szenario} | Konto: XXXX{kto}",
        ),
        ("INFO", "Alert {id} geschlossen: FALSE_POSITIVE | Bearbeiter: {user}"),
        ("ERROR", "Datenbankverbindung unterbrochen: Retry 1/3 | {ts}"),
        ("INFO", "Datenbankverbindung wiederhergestellt"),
    ]
    szenarios = ["TM-01", "TM-02", "TM-03", "TM-04", "TM-06", "TM-07"]
    users = ["s.becker", "t.mueller", "extern.zeitarbeit"]

    for day in range(1, 32):
        ts_base = f"2024-12-{day:02d}"
        for _ in range(rng.randint(15, 40)):
            hour = rng.randint(0, 23)
            minute = rng.randint(0, 59)
            ts = f"{ts_base} {hour:02d}:{minute:02d}:{rng.randint(0, 59):02d}"
            level, template = rng.choice(event_types)
            msg = template.format(
                depth=rng.randint(400, 950),
                count=rng.randint(8000, 25000),
                id=f"ALT-{rng.randint(100000, 999999)}",
                szenario=rng.choice(szenarios),
                kto=rng.randint(1000, 9999),
                user=rng.choice(users),
                ts=ts,
            )
            lines.append(f"{ts} [{level:5s}] {msg}")

    out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    log.info("  ✓ %s", out.name)


def create_goaml_log(out: Path):
    """goAML-Einreichungslog – stützt S05-02 (Fristen eingehalten)."""
    rng = random.Random(77)
    lines = [
        "# goAML FIU-Portal – Einreichungsprotokoll Musterbank AG",
        "# BLZ 500 123 45 | Exportiert: 2025-01-02",
        "# Zeitraum: 01.01.2024 – 31.12.2024",
        "",
        "SAR_REF              EINGANG_INTERN  EINREICHUNG_FIU  BEARBEITUNGSZEIT  STATUS          FIU_BESTÄTIGUNG",
        "-" * 100,
    ]
    base = date(2024, 1, 5)
    for i in range(89):
        days = rng.randint(0, 355)
        d_intern = base + timedelta(days=days)
        d_fiu = d_intern + timedelta(days=rng.randint(1, 4))
        delta = (d_fiu - d_intern).days
        fiu_ref = f"goAML-{rng.randint(800000, 899999)}"
        lines.append(
            f"SAR-2024-{1000 + i}      {d_intern}  {d_fiu}           {delta} Tage              EINGEREICHT     {fiu_ref}"
        )

    lines += [
        "",
        "# ZUSAMMENFASSUNG: 89 SARs eingereicht | Durchschn. Bearbeitungszeit: 3.2 Tage",
        "# Alle SARs innerhalb der gesetzlichen Frist (§ 43 GwG) eingereicht: JA",
    ]
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    log.info("  ✓ %s", out.name)


# ------------------------------------------------------------------ #
# Hauptroutine
# ------------------------------------------------------------------ #


def generate_all(output_dir: str):
    base = Path(output_dir)
    dirs = {
        "pdfs": base / "pdfs",
        "excel": base / "excel",
        "interviews": base / "interviews",
        "logs": base / "logs",
    }
    for d in dirs.values():
        d.mkdir(parents=True, exist_ok=True)

    log.info("\n📄 PDFs erzeugen …")
    create_gwb_bestellungsurkunde(dirs["pdfs"] / "01_gwb_bestellungsurkunde.pdf")
    create_risikoanalyse(dirs["pdfs"] / "02_risikoanalyse_2022.pdf")
    create_kyc_handbuch(dirs["pdfs"] / "03_kyc_cdd_handbuch.pdf")
    create_tm_konzept(dirs["pdfs"] / "04_tm_konzept.pdf")
    create_sar_verfahren(dirs["pdfs"] / "05_sar_verfahrensanweisung.pdf")
    create_schulungskonzept(dirs["pdfs"] / "06_schulungskonzept.pdf")
    create_aufbewahrung_iks(dirs["pdfs"] / "07_aufbewahrung_revision.pdf")
    create_gwb_jahresbericht(dirs["pdfs"] / "08_gwb_jahresbericht_2024.pdf")

    log.info("\n📊 Excel / CSV erzeugen …")
    create_tm_alerts_excel(dirs["excel"] / "tm_alerts_2024.xlsx")
    create_schulungsmatrix_excel(dirs["excel"] / "schulungsmatrix_2024.xlsx")
    create_sar_statistik_csv(dirs["excel"] / "sar_statistik_2024.csv")

    log.info("\n🎤 Interviews erzeugen …")
    create_interview_gwb(dirs["interviews"] / "interview_gwb_dr_mueller.json")
    create_interview_compliance(
        dirs["interviews"] / "interview_compliance_officer.json"
    )

    log.info("\n📋 Logs erzeugen …")
    create_tm_system_log(dirs["logs"] / "tm_system_audit_dez2024.log")
    create_goaml_log(dirs["logs"] / "goaml_einreichungen_2024.log")

    # Übersicht ausgeben
    total = sum(1 for _ in base.rglob("*") if _.is_file())
    log.info("\n%s", "=" * 60)
    log.info("✅ Demo-Dokumentation erzeugt: %s", base.resolve())
    log.info("   %d Dateien in %d Verzeichnissen", total, len(dirs))
    log.info("")
    log.info("Erwartetes Prüfbild (GwG-Prüfung):")
    log.info(
        "  ✅ KONFORM:        S04-01, S05-01, S05-02, S05-03, S02-01, S02-03, S02-04"
    )
    log.info(
        "  ⚠️  TEILKONFORM:   S01-01, S01-03, S02-02, S02-06, S03-02, S04-02, S06-01, S06-02"
    )
    log.info("  🔴 NICHT_KONFORM:  S01-02, S03-03, S08-02")
    log.info("  ❓ NICHT_PRÜFBAR:  S08-01 (IR-Bericht nicht vorgelegt)")
    log.info("")
    log.info("Pipeline starten:")
    log.info(
        '  python pipeline.py --input %s --institution "Musterbank AG" '
        "--regulatorik gwg --skeptiker",
        base.resolve(),
    )


def main():
    parser = argparse.ArgumentParser(
        description="FinRegAgents – Demo-Dokumenten-Generator (Musterbank AG)"
    )
    parser.add_argument(
        "--output", default="./demo", help="Ausgabeverzeichnis (default: ./demo)"
    )
    args = parser.parse_args()
    generate_all(args.output)


if __name__ == "__main__":
    main()
